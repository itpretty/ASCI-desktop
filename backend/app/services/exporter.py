"""Export search results to .xlsx, .pdf, .md formats."""

import io
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment


def _parse_results(results: list[dict]) -> tuple[list[str], list[dict]]:
    """Parse result rows and extract column headers."""
    all_fields: dict[str, None] = {}
    parsed_rows = []

    for r in results:
        data = json.loads(r["result_data"]) if isinstance(r["result_data"], str) else r["result_data"]
        for key in data:
            all_fields[key] = None
        parsed_rows.append(
            {
                "doc_id": r.get("doc_id", ""),
                "doc_title": r.get("doc_title") or r.get("filename", ""),
                **data,
            }
        )

    columns = ["doc_id", "doc_title"] + [k for k in all_fields if k not in ("doc_id", "doc_title")]
    return columns, parsed_rows


def export_xlsx(results: list[dict]) -> bytes:
    """Export results to Excel (.xlsx) format."""
    columns, rows = _parse_results(results)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header
    header_font = Font(bold=True)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, "")
            if value is None:
                value = ""
            elif not isinstance(value, (str, int, float)):
                value = str(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row_idx in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, min(len(str(cell_val)), 50))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 2

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_md(results: list[dict]) -> str:
    """Export results to Markdown table format."""
    columns, rows = _parse_results(results)

    lines = []
    # Header
    lines.append("| " + " | ".join(columns) + " |")
    lines.append("| " + " | ".join("---" for _ in columns) + " |")

    # Rows
    for row_data in rows:
        cells = []
        for col_name in columns:
            val = row_data.get(col_name, "")
            if val is None:
                val = ""
            val = str(val).replace("|", "\\|").replace("\n", " ")
            if len(val) > 80:
                val = val[:77] + "..."
            cells.append(val)
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def export_pdf(results: list[dict]) -> bytes:
    """Export results to PDF via weasyprint (HTML-to-PDF)."""
    columns, rows = _parse_results(results)

    # Build HTML
    html_parts = [
        "<!DOCTYPE html><html><head><meta charset='utf-8'>",
        "<style>",
        "body { font-family: sans-serif; font-size: 10px; margin: 20px; }",
        "table { border-collapse: collapse; width: 100%; }",
        "th, td { border: 1px solid #ccc; padding: 4px 6px; text-align: left; }",
        "th { background: #f5f5f5; font-weight: bold; }",
        "tr:nth-child(even) { background: #fafafa; }",
        "</style></head><body>",
        "<h1>ASCI-Desktop Export</h1>",
        "<table><thead><tr>",
    ]

    for col in columns:
        html_parts.append(f"<th>{col}</th>")
    html_parts.append("</tr></thead><tbody>")

    for row_data in rows:
        html_parts.append("<tr>")
        for col in columns:
            val = row_data.get(col, "")
            if val is None:
                val = ""
            val = str(val)
            if len(val) > 200:
                val = val[:197] + "..."
            html_parts.append(f"<td>{val}</td>")
        html_parts.append("</tr>")

    html_parts.append("</tbody></table></body></html>")
    html = "".join(html_parts)

    from weasyprint import HTML

    pdf_bytes = HTML(string=html).write_pdf()
    return pdf_bytes
